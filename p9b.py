"""Demonstrate python program to read the data from the spreadsheet and write the data
in to the spreadsheet"""

from openpyxl import Workbook

wb = Workbook()
sheet = wb.active
sheet.title = "Language"

lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code = ['KA', 'TS', 'TN']

sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Language"
sheet.cell(row=1, column=3).value = "Code"
sheet.cell(row=1, column=4).value = "Capital"

for i in range(2, 5):
    sheet.cell(row=i, column=1).value = state[i - 2]
    sheet.cell(row=i, column=2).value = lang[i - 2]
    sheet.cell(row=i, column=3).value = code[i - 2]
    sheet.cell(row=i, column=4).value = capital[i - 2]

wb.save("demo.xlsx")
srchCode = input("Enter state code: ")
found = False
for i in range(2, 5):
    data = sheet.cell(row=i, column=3).value
    if data == srchCode:
        print("Corresponding capital for code", srchCode, "is", sheet.cell(row=i, column=4).value, "and language is",
              sheet.cell(row=i, column=2).value)
        found = True

if not found:
    print('Invalid Code')
wb.close()
